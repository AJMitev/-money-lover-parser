import openpyxl as xl
from models.category import Category
from models.category_type import CategoryType
from models.transaction import Transaction
from models.currency import Currency
import uuid


def parse_data(sheet):
    categories = {}
    currencies = {}
    for row in range(2, sheet.max_row + 1):
        category_name = sheet.cell(row, 2).value
        transaction_amount = float(sheet.cell(row, 3).value)
        if category_name not in categories:
            category_type = CategoryType.Income if transaction_amount > 0 else CategoryType.Expense
            category = Category(category_name, category_type, uuid.uuid4())
            categories.update({category_name: category})
            continue
        transaction_note = sheet.cell(row, 4).value
        transaction_currency = sheet.cell(row, 6).value
        transaction_date = sheet.cell(row, 7).value

        if transaction_currency not in currencies:
            currencies.update(
                {transaction_currency: Currency(transaction_currency, uuid.uuid4())})

        transaction = Transaction(
            transaction_amount, transaction_currency, transaction_date, transaction_note, uuid.uuid4())
        categories[category_name].add_transaction(transaction)

    return {
        "categories": categories.values(),
        "currencies": currencies.values()
    }


def parse_report(workbook_name):
    wb = xl.load_workbook(workbook_name)
    data = wb["Money Lover Report"]

    result = parse_data(data)
    return result
